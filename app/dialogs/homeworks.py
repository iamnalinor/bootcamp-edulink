import asyncio
import contextlib
import io
import logging
import operator
import tempfile
import zipfile
from datetime import timedelta
from pathlib import Path
from typing import Any

from aiogram import Bot, types
from aiogram.enums import ContentType
from aiogram.types import BufferedInputFile
from aiogram_dialog import Dialog, DialogManager, ShowMode, Window
from aiogram_dialog.api.entities import MediaAttachment, MediaId
from aiogram_dialog.widgets.input import MessageInput
from aiogram_dialog.widgets.kbd import (
    Button,
    Next,
    ScrollingGroup,
    Select,
)
from aiogram_dialog.widgets.media import DynamicMedia
from aiogram_dialog.widgets.text import Format, Jinja
from aiogram_dialog.widgets.widget_event import ensure_event_processor
from openpyxl import Workbook
from openpyxl.styles import Font

from app.misc import BACK, get_client
from app.models import Container, Homework
from app.states import HomeworksSG
from app.utils import lazy_gettext as _
from app.widgets import Emojize, StartWithSameData

logger = logging.getLogger(__name__)


async def homeworks_getter(dialog_manager: DialogManager, **__: Any) -> dict[str, Any]:
    container_id = dialog_manager.start_data["container_id"]
    container = await Container.get(id=container_id).prefetch_related("owner")
    return {
        "container": container,
        "homeworks": await Homework.filter(container=container).prefetch_related(
            "owner"
        ),
    }


async def open_homework(
    __: types.CallbackQuery, ___: Button, manager: DialogManager, homework_id: str
) -> None:
    await manager.start(
        HomeworksSG.view, data=manager.start_data | {"homework_id": int(homework_id)}
    )


async def homework_view_getter(
    dialog_manager: DialogManager, **__: Any
) -> dict[str, Any]:
    container_id = dialog_manager.start_data["container_id"]
    container = await Container.get(id=container_id).prefetch_related("owner")
    homework_id = dialog_manager.start_data["homework_id"]
    homework = await Homework.get(id=homework_id).prefetch_related("owner", "container")

    created_at = (homework.created_at + timedelta(hours=3)).strftime(
        "%d.%m.%Y %H:%M:%S UTC+3"
    )

    return {
        "container": container,
        "homework": homework,
        "created_at": created_at,
        "media": MediaAttachment(
            ContentType.DOCUMENT, file_id=MediaId(file_id=homework.file_id)
        ),
        "has_no_mark": homework.mark is None,
    }


async def download_homework_all(
    call: types.CallbackQuery, __: Button, manager: DialogManager
) -> None:
    container_id = manager.start_data["container_id"]
    homeworks = await Homework.filter(container_id=container_id)
    bot: Bot = manager.middleware_data["bot"]

    await call.answer(
        _("Выгружаем решения. Это займёт некоторое время"), show_alert=True
    )

    zipio = io.BytesIO()

    with tempfile.TemporaryDirectory() as tmpdir:
        await asyncio.wait(
            [
                asyncio.create_task(
                    bot.download(
                        homework.file_id, f"{tmpdir}/{homework.id:04d}-{homework.name}"
                    )
                )
                for homework in homeworks
            ]
        )

        with zipfile.ZipFile(zipio, "w") as zipf:
            for file in Path(tmpdir).glob("*"):
                zipf.write(file, file.name)

    zipio.seek(0)
    zipio.name = f"container_{container_id}_files.zip"
    await get_client().send_document(chat_id=call.from_user.id, document=zipio)

    await manager.show(ShowMode.SEND)


async def download_homework(
    call: types.CallbackQuery,
    __: Button,
    manager: DialogManager,
) -> None:
    homework_id = manager.start_data["homework_id"]
    homework = await Homework.get(id=homework_id)
    await call.message.answer_document(homework.file_id)
    await manager.show(ShowMode.SEND)


@ensure_event_processor
async def add_mark_handler(
    message: types.Message,
    __: MessageInput,
    manager: DialogManager,
) -> None:
    mark = message.text
    if not mark.isdigit():
        await message.answer(_("Оценка должна быть целым неотрицательным числом."))
        return

    homework_id = manager.start_data["homework_id"]
    await Homework.filter(id=homework_id).update(mark=int(mark))
    await manager.done()


async def add_empty_mark(
    call: types.CallbackQuery,
    __: Button,
    manager: DialogManager,
) -> None:
    homework_id = manager.start_data["homework_id"]
    await Homework.filter(id=homework_id).update(mark=-1)
    await manager.done()


async def download_table(
    call: types.CallbackQuery, __: Button, manager: DialogManager
) -> None:
    container_id = manager.start_data["container_id"]
    container = await Container.get(id=container_id)
    homeworks = await Homework.filter(container=container).prefetch_related("owner")

    wb = Workbook()
    ws = wb.active

    headers = [
        _("ID"),
        _("Время отправки"),
        _("ФИО"),
        _("Имя файла"),
        _("Оценка"),
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = str(header)
        cell.font = Font(bold=True)

    for row, homework in enumerate(homeworks, 2):
        submitted_at = (homework.created_at + timedelta(hours=3)).strftime(
            "%d.%m.%Y %H:%M:%S"
        )
        grade = (
            _("Незачёт")
            if homework.mark == -1
            else str(homework.mark)
            if homework.mark is not None
            else ""
        )
        row_data = [
            homework.id,
            submitted_at,
            homework.owner.fio,
            homework.name,
            str(grade),
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)

    # adjust column widths
    for column_iter in ws.columns:
        column = list(column_iter)
        max_length = 0
        for cell in column:
            with contextlib.suppress(Exception):
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    excel_file = io.BytesIO()
    wb.save(excel_file)

    await call.message.answer_document(
        BufferedInputFile(
            file=excel_file.getvalue(),
            filename=f"container_{container.id}_grades.xlsx",
        ),
    )
    await manager.show(ShowMode.SEND)


homeworks_dialog = Dialog(
    Window(
        Emojize(_(":package: <b>Доступные задания</b>")),
        Button(
            Emojize(_(":inbox_tray: Скачать файлы")),
            "download_homework_all",
            download_homework_all,
        ),
        Button(
            Emojize(_(":hash: Скачать таблицу")),
            "download_table",
            download_table,
        ),
        ScrollingGroup(
            Select(
                Format("{item.owner.fio}"),
                id="s_homeworks",
                item_id_getter=operator.attrgetter("id"),
                items="homeworks",
                on_click=open_homework,
            ),
            id="scroll_homeworks",
            width=1,
            height=10,
            hide_on_single_page=True,
        ),
        BACK,
        state=HomeworksSG.intro,
        getter=homeworks_getter,
        preview_add_transitions=[Next()],
    ),
    Window(
        Emojize(
            Jinja(
                _("""
:package: <b>Контейнер {{ container.name }} › Решение {{ homework.id }}</b>
Отправил {{ homework.owner.fio }} в {{ created_at }}

{% if homework.text %}
<blockquote expandable>{{ homework.text }}</blockquote>

{% endif %}
{% if homework.mark == -1 %}
:no_entry: <b>Выставлен незачёт</b>
{% elif homework.mark %}
:hash: <b>Выставлена оценка {{ homework.mark }}</b>
{% endif %}
""")
            )
        ),
        DynamicMedia("media"),
        # Button(
        #     Emojize(_(":inbox_tray: Скачать")),
        #     "download_homework",
        #     download_homework,
        # ),
        StartWithSameData(
            Emojize(_(":hash: Выставить оценку")),
            "add_mark",
            HomeworksSG.add_mark,
            when="has_no_mark",
        ),
        BACK,
        getter=homework_view_getter,
        state=HomeworksSG.view,
    ),
    Window(
        Emojize(_(":hash: Введи оценку целым числом:")),
        MessageInput(add_mark_handler, ContentType.TEXT),
        Button(
            Emojize(_(":no_entry: Незачёт")),
            "add_mark",
            add_empty_mark,
        ),
        BACK,
        state=HomeworksSG.add_mark,
    ),
)
